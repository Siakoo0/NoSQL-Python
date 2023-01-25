from openpyxl import Workbook

from pathlib import Path

class ResultWriter:
    def __init__(self, dimension : int, name : str) -> None:
        self.dimension = dimension
        self.path = f"./data/{dimension}/benchmark"
        
        Path(self.path).mkdir(0o777, True, True)
        
        self.name = name
        self.wb = Workbook()
        
    def writeResults(self, percentage : str, rows : list):        
        columnIndex = int(percentage) / 25
        
        sheetNames = [
            "first_query",
            "second_query",
            "third_query",
            "fourth_query",
        ]
        
        for sheetName in sheetNames:
            if sheetName not in self.wb.sheetnames:
                self.wb.create_sheet(sheetName)
                
            sheet = self.wb.get_sheet_by_name(sheetName)
            sheet.cell(row=1, column=columnIndex).value = str(percentage) + "%"
        
        for indexRow, rowValue in enumerate(rows):
            for column, columnValue in enumerate(rowValue):
                sheetName = sheetNames[column % len(sheetNames)]
                
                sheet = self.wb.get_sheet_by_name(sheetName)
                
                sheet.cell(row=indexRow+2, column=columnIndex).value = columnValue
        
        
    def save(self):
        if "Sheet" in self.wb.get_sheet_names():
            del self.wb["Sheet"]
            
        self.wb.save(self.path + f"/{self.name}.xlsx")
                